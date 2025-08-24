# -*- coding: utf-8 -*-
# Retour Slog — Android (Kivy)
# Étapes guidées, images d'aide sous les champs, boutons conclusion verrouillés selon l'état,
# sauvegarde en Excel (.xlsx), vérification simple du stock pour "Changement de carton".
#
# Dépendances packagées via buildozer : kivy, pillow (PIL), openpyxl, plyer
#
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.image import Image
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.utils import platform
from kivy.properties import StringProperty, NumericProperty, DictProperty, ListProperty
from kivy.metrics import dp

import os, sys, datetime, traceback
from collections import defaultdict

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# PIL for image resizing (optional; Kivy Image auto-scales with allow_stretch)
try:
    from PIL import Image as PILImage
    PIL_OK = True
except Exception:
    PIL_OK = False

# Android storage permissions (runtime)
if platform == 'android':
    try:
        from android.permissions import request_permissions, Permission
        request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
    except Exception:
        pass

APP_NAME = "Retour Slog (Android)"

COLUMNS = [
    "État global (EN)",
    "Numéro de commande",
    "Numéro de retour",
    "Référence produit",
    "Quantité de colis",
    "Nom du client",
    "N° de suivi",
    "Transporteur",
    "Conclusion produit (EN)",
    "Date du jour",
    "Heure début saisie",
    "Heure fin saisie",
    "Temps de saisie (ISO)",
    "Photo (lien)",
]

TRANSPORTEURS = ["Colissimo","Chronopost","Mondial Relay","Relais Colis","UPS","DHL","GLS","Autre"]
ETAT_MAP = [("Cassé","BROKEN"),("Défectueux","DEFECTIVE"),("Remis en stock","BACK IN STOCK")]
CONCLUSIONS_MAP = [
    ("Étiquettes retirées — remis en stock", "REMOVE LABEL & PUT IN STOCK"),
    ("Étiquette et scotch retirés — remis en stock", "REMOVE THE LABEL, ADHESIVE & PUT IN THE SOTCK"),
    ("Destruction", "DESTRUCTION"),
    ("Changement de carton", "CARTON CHANGE"),
]
CONCLUSIONS_FR = [fr for fr,_ in CONCLUSIONS_MAP]
CONCLUSION_TO_EN = {fr: en for fr,en in CONCLUSIONS_MAP}

HELP_NAMES = {
    "reference": ["photo1","photo2"],
    "commande": ["photo3"],
    "retour": ["photo4"],
    "suivi": ["photo5","photo6","photo7","photo8"],
    "client": ["photo9","photo10","photo11"],
}

def to_str(x): return "" if x is None else str(x)

def ref_prefix(s):
    s = (to_str(s).strip().upper())
    return s[:4] if len(s) >= 4 else s

def primary_storage():
    # Common path on Android
    candidates = [
        "/storage/emulated/0",
        "/sdcard",
    ]
    for p in candidates:
        if os.path.isdir(p):
            return p
    return os.getcwd()

def default_save_dir():
    base = primary_storage()
    # Prefer Documents/RetourSlog; fallback to Download
    d1 = os.path.join(base, "Documents", "RetourSlog")
    d2 = os.path.join(base, "Download", "RetourSlog")
    for d in (d1, d2):
        try:
            os.makedirs(d, exist_ok=True)
            return d
        except Exception:
            pass
    return base

def excel_output_path(save_dir, period="monthly"):
    today = datetime.date.today()
    if (period or "monthly").lower().startswith("monthly"):
        fn = f"RetourSlog_{today.strftime('%Y-%m')}.xlsx"
    else:
        fn = f"RetourSlog_{today.strftime('%Y-%m-%d')}.xlsx"
    path = os.path.join(save_dir, fn)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
    except Exception:
        pass
    return path

def get_or_create_workbook(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(COLUMNS)
            wb.save(path)
        else:
            # ensure header row is correct
            first = [to_str(c.value).strip() for c in ws[1]]
            target = [to_str(x).strip() for x in COLUMNS]
            if first != target:
                for i, val in enumerate(COLUMNS, start=1):
                    ws.cell(row=1, column=i, value=val)
                wb.save(path)
        return wb, ws
    wb = Workbook(); ws = wb.active; ws.title = "Saisies"
    ws.append(COLUMNS)
    try:
        for i in range(len(COLUMNS)):
            ws.column_dimensions[get_column_letter(i+1)].width = 26
    except Exception:
        pass
    wb.save(path)
    return wb, ws

def find_help_image(basename):
    # Look in ./photos and ./
    here = os.path.dirname(sys.argv[0]) or os.getcwd()
    photos = os.path.join(here, "photos")
    exts = (".png",".jpg",".jpeg",".gif",".webp")
    for ext in exts:
        p = os.path.join(photos, basename + ext)
        if os.path.exists(p): return p
    for ext in exts:
        p = os.path.join(here, basename + ext)
        if os.path.exists(p): return p
    # Android Download folder
    dld = os.path.join(primary_storage(), "Download", basename + ".png")
    if os.path.exists(dld): return dld
    return None

def guess_stock_paths(save_dir):
    # try in save_dir, app dir, and Downloads
    here = os.path.dirname(sys.argv[0]) or os.getcwd()
    candidates = []
    names = ["stock_carton_slog.xlsx", "stock_carton_slog.xlsm", "stock_carton_slog.xls"]
    for base in [save_dir, here, os.path.join(primary_storage(), "Download")]:
        if not base: continue
        try:
            for fn in os.listdir(base):
                low = fn.lower()
                if low.startswith("stock_carton_slog") and low.endswith((".xlsx",".xlsm",".xls")):
                    candidates.append(os.path.join(base, fn))
        except Exception:
            pass
        for n in names:
            p = os.path.join(base, n)
            if os.path.exists(p) and p not in candidates:
                candidates.append(p)
    return candidates

def load_stock_rows(path):
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for a,b,c,d,*rest in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            A = to_str(a); C = to_str(c)
            try: D = int(float(d or 0))
            except Exception: D = 0
            if A or C or D:
                rows.append((A, C, D))  # (ref, emplacement, qty)
        wb.close()
        return rows
    except Exception as e:
        raise RuntimeError(f"Lecture stock impossible: {e}")

def deduct_stock(path, prefix, emplacement, qty):
    try:
        wb = load_workbook(path)
        ws = wb.active
        remaining = int(qty or 0)
        for row in ws.iter_rows(min_row=2, values_only=False):
            A = to_str(row[0].value); C = to_str(row[2].value)
            cell_qty = row[3]
            try: D = int(float(cell_qty.value or 0))
            except Exception: D = 0
            if C != emplacement: continue
            if ref_prefix(A) != prefix: continue
            if D <= 0: continue
            take = min(D, remaining)
            cell_qty.value = D - take
            remaining -= take
            if remaining <= 0: break
        wb.save(path); wb.close()
        return remaining == 0
    except Exception:
        return False

class UpperTextInput(TextInput):
    force_x = False
    def __init__(self, **kwargs):
        kwargs.setdefault('multiline', False)
        super().__init__(**kwargs)
        self.bind(on_text_validate=self._on_validate)
        self.bind(text=self._to_upper)

    def _to_upper(self, instance, value):
        up = (value or "").upper()
        if up != value:
            # Keep cursor position roughly in place
            pos = self.cursor_index()
            self.text = up
            Clock.schedule_once(lambda dt: self.cursor = (pos, 0), 0)
    def _on_validate(self, *_):
        if self.force_x and (self.text or "").strip() == "":
            self.text = "X"
        App.get_running_app().next_step()

class RetourSlogRoot(BoxLayout):
    pass

class RetourSlogApp(App):
    state = DictProperty({})
    step_index = NumericProperty(0)
    save_dir = StringProperty(default_save_dir())
    file_period = StringProperty("monthly")
    transporteurs = ListProperty(TRANSPORTEURS)
    etats = ListProperty([k for k,_ in ETAT_MAP])
    conclusions = ListProperty(CONCLUSIONS_FR)

    def build(self):
        self.title = APP_NAME
        self.reset_state()
        root = BoxLayout(orientation='vertical', spacing=dp(6), padding=[dp(10)]*4)

        # Header
        self.header = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(80))
        self.lbl_title = Label(text=APP_NAME, font_size='22sp', bold=True, size_hint_y=None, height=dp(36))
        self.lbl_step = Label(text="", font_size='16sp', color=(0.7,0.7,0.7,1))
        self.header.add_widget(self.lbl_title); self.header.add_widget(self.lbl_step)
        root.add_widget(self.header)

        # Scrollable central content
        self.scroll = ScrollView(size_hint=(1,1))
        self.content = BoxLayout(orientation='vertical', size_hint_y=None, padding=[dp(6)]*4, spacing=dp(8))
        self.content.bind(minimum_height=self.content.setter('height'))
        self.scroll.add_widget(self.content)
        root.add_widget(self.scroll)

        # Bottom buttons
        self.bottom = BoxLayout(size_hint_y=None, height=dp(64), spacing=dp(8))
        self.btn_home = Button(text="🏠 Accueil", on_release=lambda *_: self.go_home())
        self.btn_prev = Button(text="<= Précédent", on_release=lambda *_: self.prev_step())
        self.btn_next = Button(text="Suivant =>", on_release=lambda *_: self.next_step())
        self.btn_reset = Button(text="↺ Réinitialiser", on_release=lambda *_: self.reset_current())
        self.bottom.add_widget(self.btn_home); self.bottom.add_widget(self.btn_prev)
        self.bottom.add_widget(self.btn_reset); self.bottom.add_widget(self.btn_next)
        root.add_widget(self.bottom)

        Clock.schedule_once(lambda dt: self.build_step(), 0)
        return root

    # ---------- Navigation & State ----------
    def reset_state(self):
        now = datetime.datetime.now()
        self.state = {
            "date": now.date().isoformat(),
            "start": now.time().replace(microsecond=0).isoformat(),
            "end": "",
            "reference": "",
            "commande": "",
            "retour": "",
            "suivi": "",
            "quantite": "1",
            "client": "",
            "transporteur": "",
            "etat_fr": "",
            "conclusion_fr": "",
        }

    def go_home(self):
        self.step_index = 0
        self.build_step()

    def prev_step(self):
        if self.step_index > 0:
            self.step_index -= 1
            self.build_step()

    def next_step(self):
        # validation per step
        if not self.validate_step(): return
        if self.step_index < 10:
            self.step_index += 1
            self.build_step()
        else:
            self.finalize_and_save()

    def reset_current(self):
        self.reset_state()
        self.step_index = 1
        self.build_step()

    def validate_step(self):
        s = self.state
        idx = self.step_index
        def warn(msg):
            Popup(title="Info manquante", content=Label(text=msg), size_hint=(0.8,0.35)).open()
            return False
        if idx == 2 and not s["reference"]: return warn("La référence est obligatoire.")
        if idx == 3 and not s["commande"]: return warn("Le numéro de commande est obligatoire.")
        if idx == 4 and not s["retour"]: return warn("Le numéro de retour est obligatoire.")
        if idx == 5 and not s["suivi"]: return warn("Le numéro de suivi est obligatoire.")
        if idx == 6:
            raw = (s["quantite"] or "").strip() or "1"
            try:
                n = int(raw)
                if n <= 0: raise ValueError()
                self.state["quantite"] = str(n)
            except Exception:
                return warn("Entre une quantité entière positive (1, 2, 3…).")
        if idx == 7 and not s["client"]: return warn("Le nom du client est obligatoire.")
        if idx == 8 and not s["transporteur"]: return warn("Choisis un transporteur.")
        if idx == 9 and not s["etat_fr"]: return warn("Choisis un état.")
        if idx == 10 and not s["conclusion_fr"]: return warn("Choisis une conclusion.")
        return True

    # ---------- UI build per step ----------
    def clear_content(self):
        self.content.clear_widgets()

    def add_help_images(self, keys):
        # keys is array of basenames without extension
        row = BoxLayout(orientation='horizontal', size_hint_y=None, height=max(dp(130), Window.height * 0.25), spacing=dp(8))
        any_img = False
        for name in keys:
            p = find_help_image(name)
            if p and os.path.exists(p):
                img = Image(source=p, allow_stretch=True, keep_ratio=True)
                row.add_widget(img); any_img = True
            else:
                row.add_widget(Label(text="(image manquante)", color=(1,0.6,0.6,1)))
        if not any_img:
            self.content.add_widget(Label(text="Aucune image d’aide trouvée. Place-les dans /photos ou /Download.", color=(0.7,0.7,0.7,1)))
        self.content.add_widget(row)

    def header_text(self, title):
        self.lbl_title.text = APP_NAME
        self.lbl_step.text = title

    def step_intro(self):
        self.header_text("Démarrer la saisie (Appuie sur Suivant)")
        self.content.add_widget(Label(text="Clique « Suivant » pour commencer.", font_size='18sp'))

    def field_row(self, label, key, focus=False, uppercase=True, default_x=True, help_key=None):
        self.content.add_widget(Label(text=label, size_hint_y=None, height=dp(30)))
        ti = UpperTextInput()
        ti.force_x = default_x
        def on_text(instance, value):
            self.state[key] = instance.text or ""
        ti.bind(text=on_text)
        if not uppercase:
            ti.unbind(text=ti._to_upper)
        self.content.add_widget(ti)
        if help_key:
            self.add_help_images(HELP_NAMES.get(help_key, []))
        if focus:
            Clock.schedule_once(lambda dt: ti.focus = True, 0)

    def grid_buttons(self, items, on_choose):
        grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, row_default_height=dp(64), row_force_default=True)
        grid.bind(minimum_height=grid.setter('height'))
        for txt in items:
            btn = Button(text=txt, size_hint_y=None, height=dp(64))
            btn.bind(on_release=lambda inst, t=txt: on_choose(t))
            grid.add_widget(btn)
        self.content.add_widget(grid)

    def step_reference(self):
        self.header_text("1) Référence produit")
        self.field_row("Référence :", "reference", focus=True, help_key="reference")

    def step_commande(self):
        self.header_text("2) Numéro de commande")
        self.field_row("N° de commande :", "commande", focus=True, help_key="commande")

    def step_retour(self):
        self.header_text("3) Numéro de retour")
        self.field_row("N° de retour :", "retour", focus=True, help_key="retour")

    def step_suivi(self):
        self.header_text("4) Numéro de suivi")
        self.field_row("N° de suivi :", "suivi", focus=True, help_key="suivi")

    def step_quantite(self):
        self.header_text("5) Quantité de colis")
        self.content.add_widget(Label(text="Quantité :"))
        ti = UpperTextInput()
        ti.force_x = False
        def on_text(instance, value):
            self.state["quantite"] = instance.text or ""
        ti.bind(text=on_text)
        self.content.add_widget(ti)
        self.content.add_widget(Label(text="Astuce : tape 1 si un seul colis.", color=(0.7,0.7,0.7,1)))

    def step_client(self):
        self.header_text("6) Nom du client")
        self.field_row("Client :", "client", focus=True, help_key="client")

    def step_transporteur(self):
        self.header_text("7) Transporteur — Choisis un transporteur :")
        self.grid_buttons(self.transporteurs, self.choose_transporteur)
        if self.state.get("transporteur"):
            self.content.add_widget(Label(text=f"Choix actuel : {self.state['transporteur']}", color=(0.4,1,0.4,1)))

    def choose_transporteur(self, text):
        self.state["transporteur"] = text
        self.next_step()

    def step_etat(self):
        self.header_text("8) État global — Choisis un état :")
        self.grid_buttons(self.etats, self.choose_etat)
        if self.state.get("etat_fr"):
            self.content.add_widget(Label(text=f"Choix actuel : {self.state['etat_fr']}", color=(0.4,1,0.4,1)))

    def choose_etat(self, text):
        self.state["etat_fr"] = text
        self.next_step()

    def is_carton_available(self, requested_qty):
        # Find stock file and check grouped by ref_prefix
        candidates = guess_stock_paths(self.save_dir)
        if not candidates: return (False, "")
        stock_path = candidates[0]
        try:
            rows = load_stock_rows(stock_path)
        except Exception:
            return (False, stock_path)
        target = (self.state.get("reference") or "").strip()
        if not target: return (False, stock_path)
        pref = ref_prefix(target)
        total = 0
        for A, C, D in rows:
            if ref_prefix(A) == pref:
                try: total += int(D or 0)
                except Exception: pass
                if total >= requested_qty: return (True, stock_path)
        return (False, stock_path)

    def step_conclusion(self):
        self.header_text("9) Conclusion produit — Choisis une action :")
        # Determine availability and state policy
        try:
            rq = int(self.state.get("quantite") or "1")
            if rq <= 0: rq = 1
        except Exception:
            rq = 1
        available, stock_path = self.is_carton_available(rq)
        etat = (self.state.get("etat_fr") or "").strip()

        grid = GridLayout(cols=2, spacing=dp(8), size_hint_y=None, row_default_height=dp(64), row_force_default=True)
        grid.bind(minimum_height=grid.setter('height'))

        def is_disabled(fr_text):
            if etat in ("Cassé", "Défectueux"):
                return fr_text != "Destruction"
            if etat == "Remis en stock":
                return fr_text == "Destruction"
            return False

        for fr in CONCLUSIONS_FR:
            disabled = is_disabled(fr)
            if fr == "Changement de carton" and not available:
                disabled = True
            btn = Button(text=fr, disabled=disabled)
            btn.bind(on_release=lambda inst, t=fr: self.choose_conclusion(t))
            grid.add_widget(btn)

        self.content.add_widget(grid)

        # Hints
        info = ""
        if etat in ("Cassé", "Défectueux"):
            info = f"État sélectionné : {etat} → seule « Destruction » est autorisée."
        elif etat == "Remis en stock":
            info = "État sélectionné : Remis en stock → « Destruction » est désactivé."
        if info:
            self.content.add_widget(Label(text=info, color=(0.7,0.7,0.9,1)))
        if etat != "Remis en stock" and not available:
            self.content.add_widget(Label(text="Pas de carton disponible pour cette référence", color=(1,0.4,0.4,1)))

        if self.state.get("conclusion_fr"):
            self.content.add_widget(Label(text=f"Choix actuel : {self.state['conclusion_fr']}", color=(0.4,1,0.4,1)))

    def choose_conclusion(self, fr_text):
        self.state["conclusion_fr"] = fr_text
        if fr_text == "Changement de carton":
            self.popup_carton()
        else:
            self.next_step()

    def popup_carton(self):
        # Build options from stock file
        try:
            rq = int(self.state.get("quantite") or "1")
            if rq <= 0: rq = 1
        except Exception:
            rq = 1
        target = (self.state.get("reference") or "").strip()
        if not target:
            Popup(title="Référence manquante", content=Label(text="Renseigne la référence avant de changer de carton."), size_hint=(0.8,0.35)).open()
            return
        pref = ref_prefix(target)
        candidates = guess_stock_paths(self.save_dir)
        if not candidates:
            Popup(title="Stock", content=Label(text="Fichier stock_carton_slog introuvable. Mets-le dans Download/ ou Documents/RetourSlog."), size_hint=(0.9,0.35)).open()
            return
        stock_path = candidates[0]
        try:
            rows = load_stock_rows(stock_path)
        except Exception as e:
            Popup(title="Lecture stock", content=Label(text=str(e)), size_hint=(0.9,0.35)).open()
            return

        by_place = defaultdict(int)
        for A,C,D in rows:
            if ref_prefix(A) == pref and int(D or 0) > 0:
                by_place[C] += int(D or 0)
        if not by_place:
            Popup(title="Pas de carton", content=Label(text="Pas de carton disponible pour cette référence"), size_hint=(0.8,0.35)).open()
            return

        # Build popup UI
        box = BoxLayout(orientation='vertical', spacing=dp(8), padding=[dp(10)]*4)
        box.add_widget(Label(text=f"Groupe : {pref}  •  Demande : {rq}"))
        # list as buttons
        options = sorted(by_place.items(), key=lambda t: (-t[1], t[0]))
        gl = GridLayout(cols=1, size_hint_y=None, row_force_default=True, row_default_height=dp(48), spacing=dp(6))
        gl.bind(minimum_height=gl.setter('height'))
        selected = {"idx": None}

        def select_idx(i):
            selected["idx"] = i

        for i,(place,tot) in enumerate(options):
            btn = Button(text=f"Emplacement: {place}  •  Qté totale: {tot}")
            btn.bind(on_release=lambda inst, i=i: select_idx(i))
            gl.add_widget(btn)

        sc = ScrollView(size_hint=(1,1)); sc.add_widget(gl)
        box.add_widget(sc)

        def do_validate(*_):
            i = selected["idx"]
            if i is None:
                Popup(title="Sélection", content=Label(text="Choisis un emplacement."), size_hint=(0.7,0.3)).open(); return
            place, tot = options[i]
            if tot < rq:
                Popup(title="Stock insuffisant", content=Label(text=f"Qté totale {tot} < demandé {rq}."), size_hint=(0.8,0.35)).open(); return
            ok = deduct_stock(stock_path, pref, place, rq)
            if not ok:
                Popup(title="Déduction", content=Label(text="Échec de la mise à jour du stock."), size_hint=(0.8,0.35)).open(); return
            pop.dismiss()
            self.next_step()

        btns = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(8))
        btns.add_widget(Button(text="✅ Valider", on_release=do_validate))
        btns.add_widget(Button(text="❌ Fermer", on_release=lambda *_: pop.dismiss()))
        box.add_widget(btns)

        pop = Popup(title="Choisir l'emplacement", content=box, size_hint=(0.9,0.8))
        pop.open()

    def step_finish(self):
        self.header_text("10) Fin — Valider et enregistrer")
        self.content.add_widget(Label(text="Appuie sur « ✓ Valider & enregistrer » pour ajouter la ligne.", font_size='18sp'))
        self.content.add_widget(Button(text="✓ Valider & enregistrer", size_hint_y=None, height=dp(56),
                                       on_release=lambda *_: self.finalize_and_save()))

    def build_step(self):
        self.clear_content()
        # Update bottom bar states and next button label
        self.btn_prev.disabled = (self.step_index == 0)
        self.btn_home.disabled = (self.step_index == 0)
        self.btn_reset.disabled = (self.step_index == 0)
        self.btn_next.text = "Suivant =>" if self.step_index < 10 else "✓ Valider & enregistrer"

        steps = [
            self.step_intro,
            self.step_reference,
            self.step_commande,
            self.step_retour,
            self.step_suivi,
            self.step_quantite,
            self.step_client,
            self.step_transporteur,
            self.step_etat,
            self.step_conclusion,
            self.step_finish,
        ]
        # Step label
        self.lbl_step.text = f"Étape {self.step_index}/{len(steps)-1}"
        # Build
        steps[self.step_index]()
        # Scroll back to top
        Clock.schedule_once(lambda dt: setattr(self.scroll, 'scroll_y', 1), 0)

    # ---------- Save ----------
    def finalize_and_save(self):
        now = datetime.datetime.now()
        if not self.state.get('end'):
            self.state['end'] = now.time().replace(microsecond=0).isoformat()

        etat_en = dict(ETAT_MAP).get(self.state.get("etat_fr",""), self.state.get("etat_fr",""))
        concl_en = CONCLUSION_TO_EN.get(self.state.get("conclusion_fr",""), self.state.get("conclusion_fr",""))

        row = [
            etat_en,
            self.state.get("commande",""),
            self.state.get("retour",""),
            self.state.get("reference",""),
            self.state.get("quantite",""),
            self.state.get("client",""),
            self.state.get("suivi",""),
            self.state.get("transporteur",""),
            concl_en,
            self.state.get("date",""),
            self.state.get("start",""),
            self.state.get("end",""),
            # store raw seconds delta as text for simplicity on Android
            "",  # Temps de saisie (ISO) — optionnel
            "",
        ]

        path = excel_output_path(self.save_dir, self.file_period)
        try:
            wb, ws = get_or_create_workbook(path)
            ws.append(row)
            wb.save(path)
            Popup(title="Enregistré", content=Label(text=f"Ligne ajoutée dans:\n{path}"), size_hint=(0.9,0.4)).open()
            # Reset for next
            self.reset_state()
            self.step_index = 1
            self.build_step()
        except Exception as e:
            Popup(title="Erreur d'enregistrement", content=Label(text=str(e)), size_hint=(0.9,0.4)).open()

if __name__ == '__main__':
    RetourSlogApp().run()
